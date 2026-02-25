"""
generate_zakat_tracker.py
─────────────────────────
Generates Zakat_Tracker.xlsx from scratch.
Requires: pip install openpyxl

Run:  python generate_zakat_tracker.py
Output: Zakat_Tracker.xlsx in the current directory

All 8 sheets:
  Guide · Settings · Zakat Summary · Stocks · Cash · Debts · Ledger · Reports

All 11 features:
  1. Nisab live calculator       7. Duplicate date warning
  2. Status indicator (M)        8. Amount validation (no negatives)
  3. Ledger running total        9. Hawl tracker
  4. Year filter on Reports     10. Freeze panes
  5. Service fee summary        11. Print areas
  6. Ledger search bar
  + Column N: Brought Forward
  + AutoFilter on all data sheets
"""

from openpyxl import Workbook
from openpyxl.cell.cell import Cell as OxlCell
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.properties import PageSetupProperties

# ── Palette ──────────────────────────────────────────────────────────────────
F          = 'Aptos'
HEADER_BG  = "1F4E79";  SUBHDR_BG = "2E75B6"
INPUT_BG   = "EBF5FB";  TOTAL_BG  = "D5F5E3"
ZAKAT_BG   = "FDEBD0";  LED_BG    = "F4ECF7"
DEBT_BG    = "FADBD8";  GOLD_BG   = "FEF9E7"
BLUE   = "0000FF";  BLACK  = "000000";  GREEN  = "008000"
RED    = "C0392B";  WHITE  = "FFFFFF";  PURPLE = "6C3483"
ORANGE = "E67E22"
CURR   = '$#,##0.00;($#,##0.00);"-"'
DATEFMT= 'MM/DD/YYYY'

# Ledger rows: 1-2 search, 3 title, 4 subtitle, 5 headers, 6-205 data
LD_S = 6;   LD_E = 205
# Summary data rows
SD_S = 5;   SD_E = 14


# ── Style helpers ─────────────────────────────────────────────────────────────
def bdr(w='thin'):
    s = Side(style=w)
    return Border(left=s, right=s, top=s, bottom=s)

def th_bdr():
    m = Side(style='medium')
    return Border(left=m, right=m, top=m, bottom=m)

def hdr(cell, bg=HEADER_BG, fg=WHITE, sz=10, bold=True, wrap=True):
    cell.font      = Font(name=F, bold=bold, color=fg, size=sz)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=wrap)
    cell.border    = bdr()

def sty(cell, fg=BLACK, bg=WHITE, sz=10, bold=False,
        ha='left', va='center', wrap=False, italic=False):
    cell.font      = Font(name=F, color=fg, size=sz, bold=bold, italic=italic)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal=ha, vertical=va, wrap_text=wrap)
    cell.border    = bdr()

def force(ws, row, col, value=None):
    """Write to a cell even if openpyxl thinks it is a MergedCell."""
    nc = OxlCell(ws, row=row, column=col)
    ws._cells[(row, col)] = nc
    if value is not None:
        nc.value = value
    return nc

def wtm(ws, row, c1, c2, value, _hdr=False, **kw):
    """Write value, apply style, then merge c1:c2."""
    c = ws.cell(row=row, column=c1, value=value)
    hdr(c, **kw) if _hdr else sty(c, **kw)
    ws.merge_cells(f'{get_column_letter(c1)}{row}:{get_column_letter(c2)}{row}')
    return c

def cf(formula, fg, bg):
    ds = DifferentialStyle(
        font=Font(name=F, color=fg, bold=True),
        fill=PatternFill(start_color=bg, end_color=bg, fill_type='solid')
    )
    return Rule(type='expression', formula=[formula], dxf=ds)

def print_setup(ws, orient='landscape', title_rows=None):
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.orientation = orient
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    if title_rows:
        ws.print_title_rows = title_rows


# ══════════════════════════════════════════════════════════════════════════════
# GUIDE SHEET
# ══════════════════════════════════════════════════════════════════════════════
def build_guide(wb):
    ws = wb.create_sheet("Guide")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    for col in 'CDEFGH':
        ws.column_dimensions[col].width = 22

    # Title
    ws.row_dimensions[1].height = 32
    c = ws.cell(1, 1, "ZAKAT TRACKER — USER GUIDE & REFERENCE")
    hdr(c, bg=HEADER_BG, sz=14)
    ws.merge_cells('A1:H1')

    sections = [
        (3,  "COLOUR CODING",
             [("Blue cells",       "Manual input required"),
              ("Green cells",      "Auto-calculated — do not edit"),
              ("Red/pink cells",   "Debts or warnings"),
              ("Gold/yellow cells","Nisab-related values"),
              ("Purple cells",     "Ledger running totals")]),
        (12, "SHEET GUIDE",
             [("Guide",            "This sheet — reference only"),
              ("Settings",         "Configure payment types, services, Nisab oz values"),
              ("Zakat Summary",    "Main annual tracker — one row per Zakat year"),
              ("Stocks",           "Stock/investment account balances per year"),
              ("Cash",             "Cash and liquid asset balances per year"),
              ("Debts",            "Outstanding debt balances per year"),
              ("Ledger",           "Every individual payment — Zakat, Sadaqah, etc."),
              ("Reports",          "Filter by person and year, see breakdown by type")]),
        (24, "NISAB EXPLAINED",
             [("What is Nisab?",   "Minimum wealth threshold to owe Zakat (85g gold)"),
              ("Gold standard",    "85g ÷ 31.1035 g/oz = 2.7315 troy oz × gold price"),
              ("Silver standard",  "595g ÷ 31.1035 g/oz = 19.1358 troy oz × silver price"),
              ("Which to use?",    "Gold (2.7315 oz) is the majority scholarly default"),
              ("Hawl rule",        "Wealth must be above Nisab for one full lunar year (354 days)")]),
        (33, "HOW ZAKAT IS CALCULATED",
             [("Step 1",           "Enter your Zakat date in column A of Zakat Summary"),
              ("Step 2",           "Stocks, Cash, Debts pull automatically from their sheets"),
              ("Step 3",           "Enter current gold price (col E) and your gold oz (col F)"),
              ("Step 4",           "Net Assets (H) = Stocks + Cash + Gold − Debts"),
              ("Step 5",           "Nisab (I) = Gold Price × 2.7315 oz"),
              ("Step 6",           "Zakat Due (J) = H × 2.5% if H ≥ I, else $0"),
              ("Note",             "Zakat is on your FULL net wealth, not just the surplus above Nisab")]),
        (44, "FAQ",
             [("Add payment type", "Go to Settings → Payment Types → type in an empty slot"),
              ("Add service",      "Go to Settings → Transfer Services → type in an empty slot"),
              ("Change Nisab oz",  "Settings → Nisab Settings → edit cell D44 (Gold) or D45 (Silver)"),
              ("Check today's Nisab", "Settings → Nisab Calculator → enter today's gold price in B49"),
              ("Record a payment", "Ledger sheet → fill Date, Type, Service, Given To, Amount, Fees"),
              ("View by person",   "Reports sheet → select name from C4 dropdown"),
              ("Filter by year",   "Reports sheet → select year from C5 dropdown")]),
    ]

    for start_row, title, rows in sections:
        ws.row_dimensions[start_row].height = 24
        c = ws.cell(start_row, 1, title)
        hdr(c, bg=SUBHDR_BG, sz=10)
        ws.merge_cells(f'A{start_row}:H{start_row}')

        for i, (label, desc) in enumerate(rows):
            r = start_row + 1 + i
            ws.row_dimensions[r].height = 20
            bg = "F0F4F8" if i % 2 == 0 else WHITE
            c = ws.cell(r, 1, label)
            sty(c, fg=BLACK, bg=bg, bold=True, sz=9)
            ws.merge_cells(f'A{r}:B{r}')
            c = ws.cell(r, 3, desc)
            sty(c, fg="444444", bg=bg, sz=9)
            ws.merge_cells(f'C{r}:H{r}')

    print("  ✓ Guide")


# ══════════════════════════════════════════════════════════════════════════════
# SETTINGS SHEET
# ══════════════════════════════════════════════════════════════════════════════
def build_settings(wb):
    ws = wb.create_sheet("Settings")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 35

    # Title
    ws.row_dimensions[1].height = 32
    c = ws.cell(1, 1, "ZAKAT TRACKER — SETTINGS")
    hdr(c, bg=HEADER_BG, sz=14)
    ws.merge_cells('A1:F1')

    # ── Payment Types + Transfer Services ───────────────────────────────────
    ws.row_dimensions[3].height = 24
    c = ws.cell(3, 2, "PAYMENT TYPES")
    hdr(c, bg=SUBHDR_BG, sz=10)
    ws.merge_cells('B3:C3')

    c = ws.cell(3, 4, "TRANSFER SERVICES")
    hdr(c, bg=SUBHDR_BG, sz=10)
    ws.merge_cells('D3:E3')

    c = ws.cell(3, 6, "RECIPIENTS / GIVEN TO")
    hdr(c, bg=SUBHDR_BG, sz=10)
    ws.column_dimensions['F'].width = 28
    ws.merge_cells('F3:G3')

    ws.row_dimensions[4].height = 20
    for col, lbl in [(2, "Type Name"), (4, "Service Name"), (6, "Recipient Name")]:
        c = ws.cell(4, col, lbl)
        hdr(c, bg="3D5A80", sz=9)

    preset_types    = ["Zakat", "Sadaqah", "Fitrana", "Qurbani"]
    preset_services = ["Remitly", "Wise", "Bank Transfer", "Cash", "Zelle"]
    preset_recipients = [
        "Islamic Relief USA", "Zakat Foundation", "LaunchGood",
        "Local Mosque", "Family Member",
    ]

    for i in range(30):
        r  = 5 + i
        bg = "F0F4F8" if i % 2 == 0 else WHITE
        ws.row_dimensions[r].height = 20

        # Type slot
        val = preset_types[i] if i < len(preset_types) else ""
        c = ws.cell(r, 2, val)
        sty(c, fg=BLUE if val else BLACK, bg=INPUT_BG if val else bg, sz=10, bold=bool(val))

        # Service slot
        val2 = preset_services[i] if i < len(preset_services) else ""
        c = ws.cell(r, 4, val2)
        sty(c, fg=BLUE if val2 else BLACK, bg=INPUT_BG if val2 else bg, sz=10, bold=bool(val2))

        # Recipients slot
        val3 = preset_recipients[i] if i < len(preset_recipients) else ""
        c = ws.cell(r, 6, val3)
        sty(c, fg=BLUE if val3 else BLACK, bg=INPUT_BG if val3 else bg, sz=10, bold=bool(val3))

    # "Add more" hints
    ws.row_dimensions[35].height = 18
    c = ws.cell(35, 2, "← Add more types above")
    sty(c, fg="AAAAAA", bg=WHITE, sz=8, italic=True)
    c = ws.cell(35, 4, "← Add more services above")
    sty(c, fg="AAAAAA", bg=WHITE, sz=8, italic=True)
    c = ws.cell(35, 6, "← Add recipients above")
    sty(c, fg="AAAAAA", bg=WHITE, sz=8, italic=True)

    # ── Nisab Settings ───────────────────────────────────────────────────────
    ws.row_dimensions[37].height = 26
    c = ws.cell(37, 1, "NISAB SETTINGS")
    hdr(c, bg=HEADER_BG, sz=11)
    ws.merge_cells('A37:F37')

    nisab_exp = [
        ("What is Nisab?",
         "Nisab is the minimum threshold of wealth a Muslim must own before Zakat becomes obligatory. "
         "If your net zakatable assets are below Nisab on your Zakat date, no Zakat is due. "
         "Once above it, 2.5% of your FULL net wealth is due."),
        ("Two standards:",
         "Scholars use two measures: (1) Gold Nisab = 85 grams of gold. "
         "(2) Silver Nisab = 595 grams of silver. "
         "Gold standard (85g) is the most widely used. Silver results in a lower threshold (more people qualify)."),
        ("How it's calculated:",
         "Gold Nisab = 85g ÷ 31.1035g per troy oz = 2.7315 troy oz × current gold price. "
         "So if gold is $3,000/oz, Nisab = 2.7315 × $3,000 = $8,194.50"),
    ]
    for i, (lbl, txt) in enumerate(nisab_exp):
        r = 38 + i
        ws.row_dimensions[r].height = 40
        c = ws.cell(r, 2, lbl)
        sty(c, fg=WHITE, bg=SUBHDR_BG, sz=9, bold=True, ha='center', va='center')
        c = ws.cell(r, 3, txt)
        sty(c, fg="333333", bg="F0F4F8", sz=9, wrap=True, va='center')
        ws.merge_cells(f'C{r}:F{r}')

    # Controls header
    ws.row_dimensions[42].height = 22
    for col, lbl in [(2,"Setting"),(4,"Your Value"),(6,"Notes")]:
        c = ws.cell(42, col, lbl)
        hdr(c, bg="3D5A80", sz=9)

    nisab_controls = [
        (43, "Nisab Standard",      "Gold",    "Gold (85g) or Silver (595g) — Gold is the default"),
        (44, "Gold Nisab (troy oz)", 2.7315,    "85g ÷ 31.1035 g/oz = 2.7315 oz  |  Change if your scholar uses a different value"),
        (45, "Silver Nisab (troy oz)",19.1358,  "595g ÷ 31.1035 g/oz = 19.1358 oz  |  For reference if using silver standard"),
    ]
    for row, setting, value, note in nisab_controls:
        ws.row_dimensions[row].height = 26
        c = ws.cell(row, 2, setting)
        sty(c, fg=BLACK, bg="F8F9FA", sz=9, bold=True)
        c = ws.cell(row, 4, value)
        sty(c, fg=BLUE, bg=GOLD_BG, sz=11, bold=True, ha='center')
        c.border = th_bdr()
        if isinstance(value, float):
            c.number_format = '0.0000'
        c = ws.cell(row, 6, note)
        sty(c, fg="555555", bg="F8F9FA", sz=8, wrap=True)

    # ── Nisab Live Calculator ────────────────────────────────────────────────
    ws.row_dimensions[46].height = 8

    ws.row_dimensions[47].height = 26
    c = ws.cell(47, 1, "⚡  CURRENT NISAB CALCULATOR  —  Enter today's gold price to see threshold instantly")
    hdr(c, bg=SUBHDR_BG, sz=10)
    ws.merge_cells('A47:F47')

    ws.row_dimensions[48].height = 22
    for col, lbl in [(2,"Today's Gold Price ($/oz)"), (4,"Gold Nisab Today ($)"), (6,"Silver Price ($/oz) — optional")]:
        c = ws.cell(48, col, lbl)
        hdr(c, bg="3D5A80", sz=9)

    ws.row_dimensions[49].height = 32
    c = ws.cell(49, 2, 0)
    sty(c, fg=BLUE, bg="FFFDE7", sz=13, bold=True, ha='center')
    c.border = th_bdr(); c.number_format = '$#,##0.00'

    c = ws.cell(49, 4, '=IF(B49=0,"Enter price →",B49*$D$44)')
    sty(c, fg=GREEN, bg=TOTAL_BG, sz=13, bold=True, ha='center')
    c.border = th_bdr(); c.number_format = CURR

    c = ws.cell(49, 6, 0)
    sty(c, fg=BLUE, bg="FFFDE7", sz=11, bold=True, ha='center')
    c.border = th_bdr(); c.number_format = '$#,##0.00'

    ws.row_dimensions[50].height = 20
    c = ws.cell(50, 4, '=IF(B49=0,"","Silver Nisab = "&TEXT(F49*$D$45,"$#,##0.00"))')
    sty(c, fg="666666", bg="F8F9FA", sz=9, italic=True, ha='center')

    print("  ✓ Settings")


# ══════════════════════════════════════════════════════════════════════════════
# STOCKS / CASH / DEBTS SHEETS
# ══════════════════════════════════════════════════════════════════════════════
def build_asset_sheet(wb, name, account_cols, total_col_name):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False

    col_widths = [14] + [18]*6 + [18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    titles = {
        'Stocks': ("STOCK & INVESTMENT PORTFOLIO",
                   "Enter total value of each account on your Zakat calculation date. Total auto-sums."),
        'Cash':   ("CASH & LIQUID ASSETS",
                   "Enter total balance of each account on your Zakat calculation date. Total auto-sums."),
        'Debts':  ("OUTSTANDING DEBTS",
                   "Enter current balances of debts on your Zakat calculation date. Total auto-sums."),
    }
    title, subtitle = titles[name]
    bg = DEBT_BG if name == 'Debts' else INPUT_BG
    fg_val = RED if name == 'Debts' else BLUE

    # Row 1: title
    ws.row_dimensions[1].height = 30
    c = ws.cell(1, 1, title)
    hdr(c, bg=HEADER_BG, sz=13)
    ws.merge_cells('A1:H1')

    # Row 2: subtitle
    ws.row_dimensions[2].height = 30
    c = ws.cell(2, 1, subtitle)
    sty(c, fg="444444", bg="F0F4F8", sz=9, italic=True, ha='center', wrap=True)
    ws.merge_cells('A2:H2')

    # Row 3: headers + autofilter
    ws.row_dimensions[3].height = 26
    all_cols = ["Date"] + account_cols + [total_col_name]
    for i, h_text in enumerate(all_cols, 1):
        c = ws.cell(3, i, h_text)
        col_bg = "C0392B" if (name == 'Debts' and i == 8) else HEADER_BG
        hdr(c, bg=col_bg, sz=9)

    ws.auto_filter.ref = 'A3:H13'
    ws.freeze_panes = 'A4'

    # Data rows 4-13
    for ri in range(10):
        row = 4 + ri
        ws.row_dimensions[row].height = 22
        row_bg = "FDF2F8" if (name == 'Debts' and ri % 2 == 0) else ("F0F4F8" if ri % 2 == 0 else WHITE)

        # Date
        c = ws.cell(row, 1)
        sty(c, fg=BLUE, bg=INPUT_BG, sz=10, ha='center')
        c.number_format = DATEFMT

        # Account cols 2-7
        for ci in range(2, 8):
            c = ws.cell(row, ci)
            sty(c, fg=fg_val, bg=row_bg, sz=10, ha='right')
            c.number_format = CURR

        # Total col 8
        c = ws.cell(row, 8, f'=SUM(B{row}:G{row})')
        total_bg = DEBT_BG if name == 'Debts' else TOTAL_BG
        total_fg = RED if name == 'Debts' else GREEN
        sty(c, fg=total_fg, bg=total_bg, sz=10, bold=True, ha='right')
        c.number_format = CURR

    # Print setup
    ws.print_area = 'A1:H13'
    print_setup(ws, title_rows='1:3')

    print(f"  ✓ {name}")


# ══════════════════════════════════════════════════════════════════════════════
# ZAKAT SUMMARY SHEET
# ══════════════════════════════════════════════════════════════════════════════
def build_summary(wb):
    ws = wb.create_sheet("Zakat Summary")
    ws.sheet_view.showGridLines = False

    col_widths = [14, 18, 18, 14, 16, 16, 16, 20, 16, 16, 18, 16, 18, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Row 1
    ws.row_dimensions[1].height = 32
    c = ws.cell(1, 1, "ZAKAT CALCULATOR — ANNUAL SUMMARY")
    hdr(c, bg=HEADER_BG, sz=14)
    ws.merge_cells('A1:N1')

    # Row 2
    ws.row_dimensions[2].height = 28
    c = ws.cell(2, 1,
        "Each row = one Zakat year. Stocks, Cash, Debts auto-pull from their sheets. "
        "Enter gold price + oz. All calculations are automatic.")
    sty(c, fg="444444", bg="F0F4F8", sz=9, italic=True, ha='center', wrap=True)
    ws.merge_cells('A2:N2')

    # Row 3: headers
    ws.row_dimensions[3].height = 32
    headers = [
        "Zakat Date", "Stock Portfolio ($)", "Cash & Liquid ($)", "Total Debts ($)",
        "Gold Price ($/oz)", "Gold Owned (oz)", "Value of Gold ($)",
        "Net Zakatable Assets ($)", "Nisab\nThreshold ($)", "Zakat Due (2.5%) ($)",
        "Paid This Period ($)", "Running Balance ($)", "Status", "Brought\nForward ($)"
    ]
    hdr_bgs = [HEADER_BG]*12 + [SUBHDR_BG, "7D6608"]
    for i, (h, bg) in enumerate(zip(headers, hdr_bgs), 1):
        c = ws.cell(3, i, h)
        hdr(c, bg=bg, sz=9)

    ws.auto_filter.ref = f'A3:N{SD_E}'
    ws.freeze_panes = 'A5'

    # Row 4: explanations
    ws.row_dimensions[4].height = 36
    explanations = [
        "Enter your Zakat calculation date",
        "Auto-pulled from Stocks sheet total",
        "Auto-pulled from Cash sheet total",
        "Auto-pulled from Debts sheet total",
        "Enter gold spot price on this date ($/troy oz)",
        "Enter how many troy oz of gold you own",
        "Gold value = Price × oz (auto-calculated)",
        "Stocks + Cash + Gold − Debts",
        "Nisab = minimum wealth to owe Zakat. Calculated as gold price × 2.7315 oz (= 85g gold). "
        "Edit the oz in Settings → Nisab Settings. Zakat is due ONLY if Net Assets ≥ Nisab.",
        "2.5% of Net Assets if ≥ Nisab, else $0",
        "Zakat payments from Ledger in this period",
        "Cumulative balance: positive = still owe",
        "Payment status for this year",
        "Unpaid Zakat carried in from prior year",
    ]
    exp_bgs = ["2471A3"]*14
    for i, (txt, bg) in enumerate(zip(explanations, exp_bgs), 1):
        c = ws.cell(4, i, txt)
        sty(c, fg=WHITE, bg=bg, sz=8, italic=True, ha='center', wrap=True)

    # Data rows 5-14
    for ri in range(10):
        row = SD_S + ri
        ws.row_dimensions[row].height = 22
        prev = row - 1

        # A: Date
        c = ws.cell(row, 1)
        sty(c, fg=BLUE, bg=INPUT_BG, sz=10, ha='center')
        c.number_format = DATEFMT

        # B: Stocks (pull from Stocks sheet)
        c = ws.cell(row, 2,
            f'=IFERROR(INDEX(Stocks!$A$1:$H$13,MATCH(A{row},Stocks!$A$4:$A$13,0)+3,'
            f'MATCH("Total Portfolio",Stocks!$A$3:$H$3,0)),0)')
        sty(c, fg=GREEN, bg=TOTAL_BG, sz=10, ha='right')
        c.number_format = CURR

        # C: Cash
        c = ws.cell(row, 3,
            f'=IFERROR(INDEX(Cash!$A$1:$H$13,MATCH(A{row},Cash!$A$4:$A$13,0)+3,'
            f'MATCH("Total Cash",Cash!$A$3:$H$3,0)),0)')
        sty(c, fg=GREEN, bg=TOTAL_BG, sz=10, ha='right')
        c.number_format = CURR

        # D: Debts
        c = ws.cell(row, 4,
            f'=IFERROR(INDEX(Debts!$A$1:$H$13,MATCH(A{row},Debts!$A$4:$A$13,0)+3,'
            f'MATCH("Total Debts",Debts!$A$3:$H$3,0)),0)')
        sty(c, fg=RED, bg=DEBT_BG, sz=10, ha='right')
        c.number_format = CURR

        # E: Gold price (manual)
        c = ws.cell(row, 5)
        sty(c, fg=BLUE, bg=INPUT_BG, sz=10, ha='right')
        c.number_format = CURR

        # F: Gold oz (manual)
        c = ws.cell(row, 6)
        sty(c, fg=BLUE, bg=INPUT_BG, sz=10, ha='right')
        c.number_format = '0.0000'

        # G: Gold value
        c = ws.cell(row, 7, f'=IF(OR(E{row}="",F{row}=""),0,E{row}*F{row})')
        sty(c, fg=GREEN, bg=TOTAL_BG, sz=10, ha='right')
        c.number_format = CURR

        # H: Net assets (use 0 for any empty component)
        c = ws.cell(row, 8, f'=IF(A{row}="","",B{row}+C{row}+G{row}-D{row})')
        sty(c, fg=GREEN, bg=TOTAL_BG, sz=10, bold=True, ha='right')
        c.number_format = CURR

        # I: Nisab
        c = ws.cell(row, 9, f'=IF(OR(A{row}="",E{row}=""),0,E{row}*Settings!$D$44)')
        sty(c, fg="7D6608", bg=GOLD_BG, sz=10, ha='right')
        c.number_format = CURR

        # J: Zakat due
        c = ws.cell(row, 10, f'=IF(A{row}="","",IF(H{row}>=I{row},H{row}*0.025,0))')
        sty(c, fg=ORANGE, bg=ZAKAT_BG, sz=10, bold=True, ha='right')
        c.number_format = CURR

        # K: Paid (SUMIFS from Ledger)
        if ri == 0:
            k_formula = (
                f'=IF(A{row}="","",SUMIFS(Ledger!$H${LD_S}:$H${LD_E},'
                f'Ledger!$B${LD_S}:$B${LD_E},"Zakat",'
                f'Ledger!$A${LD_S}:$A${LD_E},"<="&A{row}))'
            )
        else:
            k_formula = (
                f'=IF(A{row}="","",SUMIFS(Ledger!$H${LD_S}:$H${LD_E},'
                f'Ledger!$B${LD_S}:$B${LD_E},"Zakat",'
                f'Ledger!$A${LD_S}:$A${LD_E},">"&A{prev},'
                f'Ledger!$A${LD_S}:$A${LD_E},"<="&A{row}))'
            )
        c = ws.cell(row, 11, k_formula)
        sty(c, fg=BLUE, bg=INPUT_BG, sz=10, ha='right')
        c.number_format = CURR

        # L: Running balance
        if ri == 0:
            l_formula = f'=IF(A{row}="","",J{row}-IF(K{row}="",0,K{row}))'
        else:
            l_formula = (
                f'=IF(A{row}="","",J{row}-IF(K{row}="",0,K{row})'
                f'+IF(OR(L{prev}="",A{prev}=""),0,L{prev}))'
            )
        c = ws.cell(row, 12, l_formula)
        sty(c, fg=PURPLE, bg=LED_BG, sz=10, bold=True, ha='right')
        c.number_format = CURR

        # M: Status
        c = ws.cell(row, 13,
            f'=IF(A{row}="","",IF(J{row}=0,"N/A",IF(L{row}<=0,"✅ Paid in Full",'
            f'IF(K{row}=0,"❌ Not Started","⚠️ Partially Paid"))))')
        sty(c, fg=BLACK, bg="F8F9FA", sz=10, bold=True, ha='center')

        # N: Brought forward
        if ri == 0:
            n_formula = f'=IF(A{row}="","",0)'
        else:
            n_formula = f'=IF(A{row}="","",IF(L{prev}="",0,MAX(0,L{prev})))'
        c = ws.cell(row, 14, n_formula)
        sty(c, fg=BLACK, bg=GOLD_BG, sz=10, bold=True, ha='right')
        c.number_format = CURR

    # Conditional formats
    ws.conditional_formatting.add('A5:A14',
        cf('AND(A5<>"",COUNTIF($A$5:$A$14,A5)>1)', RED, "FFCCCC"))

    for emoji, bg_c, fg_c in [("✅","D5F5E3",GREEN),("⚠️","FDEBD0",ORANGE),("❌","FADBD8",RED)]:
        ws.conditional_formatting.add('M5:M14',
            cf(f'NOT(ISERROR(SEARCH("{emoji}",M5)))', fg_c, bg_c))

    ws.conditional_formatting.add('N5:N14',
        cf('AND(N5<>"",N5>0)', RED, "FADBD8"))

    # ── Dashboard ────────────────────────────────────────────────────────────
    DASH = SD_E + 3   # row 17
    ws.row_dimensions[DASH-1].height = 10
    ws.row_dimensions[DASH].height   = 26

    c = ws.cell(DASH, 1, "ZAKAT SUMMARY DASHBOARD")
    hdr(c, bg=HEADER_BG, sz=11)
    ws.merge_cells(f'A{DASH}:N{DASH}')

    DASH_LBL = DASH + 1;  DASH_VAL = DASH + 2
    ws.row_dimensions[DASH_LBL].height = 36
    ws.row_dimensions[DASH_VAL].height = 40

    dash_items = [
        (1,  2,  "Total Zakat Owed\n(All Years)",
         f'=SUMIF(J{SD_S}:J{SD_E},">"&0,J{SD_S}:J{SD_E})', ZAKAT_BG, ORANGE),
        (3,  4,  "Total Zakat Paid\n(from Ledger)",
         f'=SUMIF(Ledger!$B${LD_S}:$B${LD_E},"Zakat",Ledger!$H${LD_S}:$H${LD_E})', TOTAL_BG, GREEN),
        (5,  6,  "Total Sadaqah Paid\n(from Ledger)",
         f'=SUMIF(Ledger!$B${LD_S}:$B${LD_E},"Sadaqah",Ledger!$H${LD_S}:$H${LD_E})', INPUT_BG, BLUE),
        (7,  8,  "Total Fitrana Paid\n(from Ledger)",
         f'=SUMIF(Ledger!$B${LD_S}:$B${LD_E},"Fitrana",Ledger!$H${LD_S}:$H${LD_E})', LED_BG, PURPLE),
        (9,  10, "Total Qurbani Paid\n(from Ledger)",
         f'=SUMIF(Ledger!$B${LD_S}:$B${LD_E},"Qurbani",Ledger!$H${LD_S}:$H${LD_E})', GOLD_BG, "7D6608"),
        (11, 12, "Outstanding Balance\n(Owed − Paid)",
         f'=SUMIF(J{SD_S}:J{SD_E},">"&0,J{SD_S}:J{SD_E})-SUMIF(Ledger!$B${LD_S}:$B${LD_E},"Zakat",Ledger!$H${LD_S}:$H${LD_E})',
         DEBT_BG, RED),
    ]
    for c1, c2, lbl, formula, bg, fg in dash_items:
        c = ws.cell(DASH_LBL, c1, lbl)
        hdr(c, bg=SUBHDR_BG, sz=9)
        ws.merge_cells(f'{get_column_letter(c1)}{DASH_LBL}:{get_column_letter(c2)}{DASH_LBL}')
        c = ws.cell(DASH_VAL, c1, formula)
        sty(c, fg=fg, bg=bg, sz=14, bold=True, ha='center')
        c.number_format = CURR
        c.border = th_bdr()
        ws.merge_cells(f'{get_column_letter(c1)}{DASH_VAL}:{get_column_letter(c2)}{DASH_VAL}')

    # ── Hawl Tracker ─────────────────────────────────────────────────────────
    HAWL_T = DASH_VAL + 3
    ws.row_dimensions[HAWL_T-1].height = 10
    ws.row_dimensions[HAWL_T].height   = 28

    c = ws.cell(HAWL_T, 1, "🌙  HAWL TRACKER — Next Zakat Due Date")
    hdr(c, bg=HEADER_BG, sz=11)
    ws.merge_cells(f'A{HAWL_T}:N{HAWL_T}')

    HAWL_E = HAWL_T + 1
    ws.row_dimensions[HAWL_E].height = 32
    c = ws.cell(HAWL_E, 1,
        "Hawl = one complete lunar year (354 days) of wealth above Nisab. "
        "Shows countdown to your next Zakat anniversary based on your most recent Zakat date.")
    sty(c, fg="333333", bg="F0F4F8", sz=9, italic=True, ha='center', wrap=True)
    ws.merge_cells(f'A{HAWL_E}:N{HAWL_E}')

    HAWL_H = HAWL_E + 1;  HAWL_V = HAWL_H + 1
    ws.row_dimensions[HAWL_H].height = 26
    ws.row_dimensions[HAWL_V].height = 38

    hawl_cols = [(1,2,"Last Zakat Date"),(3,4,"Next Due (+354 days)"),
                 (5,6,"Days Remaining"),(7,9,"Status"),(10,11,"Today (auto)")]
    for c1, c2, lbl in hawl_cols:
        cell = force(ws, HAWL_H, c1, lbl)
        hdr(cell, bg=SUBHDR_BG, sz=9)
        if c1 != c2:
            ws.merge_cells(f'{get_column_letter(c1)}{HAWL_H}:{get_column_letter(c2)}{HAWL_H}')
            for cx in range(c1+1, c2+1):
                force(ws, HAWL_H, cx)

    hawl_vals = [
        (1, 2, f'=IF(COUNTA(A{SD_S}:A{SD_E})=0,"No dates yet",MAX(A{SD_S}:A{SD_E}))',
         BLUE, INPUT_BG, DATEFMT),
        (3, 4, f'=IF(A{HAWL_V}="No dates yet","",A{HAWL_V}+354)',
         PURPLE, LED_BG, DATEFMT),
        (5, 6, f'=IF(C{HAWL_V}="","",MAX(0,C{HAWL_V}-TODAY()))',
         BLACK, GOLD_BG, '0 "days"'),
        (7, 9, f'=IF(C{HAWL_V}="","",IF(TODAY()>C{HAWL_V},"🕌 Zakat Due Now!",'
               f'IF(C{HAWL_V}-TODAY()<=30,"⚠️ Due Soon (< 30 days)","✅ Hawl in progress")))',
         BLACK, "F8F9FA", None),
        (10,11,'=TODAY()', "888888", "F2F3F4", DATEFMT),
    ]
    for c1, c2, formula, fg, bg, fmt in hawl_vals:
        cell = force(ws, HAWL_V, c1, formula)
        sty(cell, fg=fg, bg=bg, sz=11, bold=True, ha='center')
        if fmt: cell.number_format = fmt
        if c1 != c2:
            ws.merge_cells(f'{get_column_letter(c1)}{HAWL_V}:{get_column_letter(c2)}{HAWL_V}')
            for cx in range(c1+1, c2+1):
                force(ws, HAWL_V, cx)

    for emoji, bg_c, fg_c in [("🕌","FADBD8",RED),("⚠️","FDEBD0",ORANGE),("✅","D5F5E3",GREEN)]:
        ws.conditional_formatting.add(f'G{HAWL_V}:I{HAWL_V}',
            cf(f'NOT(ISERROR(SEARCH("{emoji}",G{HAWL_V})))', fg_c, bg_c))

    # Print setup
    ws.print_area = f'A1:N{HAWL_V}'
    print_setup(ws, title_rows='1:4')

    print("  ✓ Zakat Summary")


# ══════════════════════════════════════════════════════════════════════════════
# LEDGER SHEET
# ══════════════════════════════════════════════════════════════════════════════
def build_ledger(wb):
    ws = wb.create_sheet("Ledger")
    ws.sheet_view.showGridLines = False

    col_widths = [13, 14, 16, 18, 28, 14, 12, 15, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Rows 1-2: Search bar
    ws.row_dimensions[1].height = 30
    c = ws.cell(1, 1, "🔍  Search / Filter:")
    sty(c, fg=WHITE, bg=SUBHDR_BG, sz=10, bold=True, ha='right')
    ws.merge_cells('A1:B1')

    c = ws.cell(1, 3, "")
    sty(c, fg=BLUE, bg="FFFDE7", sz=11, bold=True, ha='left')
    c.border = th_bdr()

    c = ws.cell(1, 4, "Type a name, keyword, or type — matching rows highlight yellow. Leave blank to show all.")
    sty(c, fg="555555", bg="F8F9FA", sz=9, italic=True, ha='left', wrap=True)
    ws.merge_cells('D1:I1')

    ws.row_dimensions[2].height = 20
    c = ws.cell(2, 1,
        f'=IF(C1="","All entries shown",'
        f'COUNTIF(D{LD_S}:D{LD_E},"*"&C1&"*")+'
        f'COUNTIF(E{LD_S}:E{LD_E},"*"&C1&"*")+'
        f'COUNTIF(B{LD_S}:B{LD_E},"*"&C1&"*")&" matching rows highlighted")')
    sty(c, fg=GREEN, bg=TOTAL_BG, sz=9, bold=True, ha='center')
    ws.merge_cells('A2:I2')

    # Row 3: Title (merged and centered)
    ws.row_dimensions[3].height = 30
    c = ws.cell(3, 1, "ZAKAT & SADAQAH LEDGER")
    hdr(c, bg=HEADER_BG, sz=13)
    ws.merge_cells('A3:I3')

    # Row 4: Subtitle
    ws.row_dimensions[4].height = 28
    c = ws.cell(4, 1,
        "Record every payment made — Zakat, Sadaqah, Fitrana, Qurbani. "
        "Amount + Fees = Total Paid. Use the Type & Service dropdowns.")
    sty(c, fg="444444", bg="F0F4F8", sz=9, italic=True, ha='center', wrap=True)
    ws.merge_cells('A4:I4')

    # Row 5: Headers + autofilter
    ws.row_dimensions[5].height = 26
    headers = ["Date","Type","Service Used","Given To","Details / Notes",
               "Amount ($)","Fees ($)","Total Paid ($)","Running Total ($)"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(5, i, h)
        hdr(c, bg=SUBHDR_BG if i == 9 else HEADER_BG, sz=9)

    ws.auto_filter.ref = f'A5:I{LD_E}'
    ws.freeze_panes = 'A6'

    # Type + Service dropdowns
    type_dv = DataValidation(type="list", formula1="Settings!$B$5:$B$34",
                              allow_blank=True, showDropDown=False)
    type_dv.error = "Please select a type from the list, or add it in Settings."
    type_dv.errorTitle = "Invalid Type"
    ws.add_data_validation(type_dv)
    type_dv.sqref = f'B{LD_S}:B{LD_E}'

    svc_dv = DataValidation(type="list", formula1="Settings!$D$5:$D$34",
                             allow_blank=True, showDropDown=False)
    svc_dv.error = "Please select a service from the list, or add it in Settings."
    svc_dv.errorTitle = "Invalid Service"
    ws.add_data_validation(svc_dv)
    svc_dv.sqref = f'C{LD_S}:C{LD_E}'

    # Given To dropdown (recipients from Settings col F)
    given_dv = DataValidation(type="list", formula1="Settings!$F$5:$F$34",
                              allow_blank=True, showDropDown=False)
    given_dv.error      = "Select a recipient or add them in Settings → Recipients / Given To."
    given_dv.errorTitle = "Unknown Recipient"
    given_dv.prompt     = "Pick from list or type a name"
    given_dv.promptTitle= "Given To"
    ws.add_data_validation(given_dv)
    given_dv.sqref = f'D{LD_S}:D{LD_E}'

    # Amount validation (no negatives)
    amt_dv = DataValidation(type="decimal", operator="greaterThanOrEqual",
                             formula1="0", allow_blank=True)
    amt_dv.error = "Amount must be 0 or greater."
    amt_dv.errorTitle = "Invalid Amount"
    ws.add_data_validation(amt_dv)
    amt_dv.sqref = f'F{LD_S}:G{LD_E}'

    # Data rows 6-205
    for i in range(200):
        row = LD_S + i
        bg  = "F0EAF8" if i % 2 == 0 else "E8DCF5"
        ws.row_dimensions[row].height = 20

        c = ws.cell(row, 1)
        sty(c, fg=BLUE, bg=INPUT_BG, sz=9, ha='center')
        c.number_format = DATEFMT

        for col in [2, 3, 4, 5]:
            c = ws.cell(row, col)
            sty(c, fg=BLACK, bg=bg, sz=9)

        for col in [6, 7]:
            c = ws.cell(row, col)
            sty(c, fg=BLUE, bg=INPUT_BG, sz=9, ha='right')
            c.number_format = CURR

        # H: Total Paid
        c = ws.cell(row, 8, f'=IF(F{row}+G{row}=0,"",F{row}+G{row})')
        sty(c, fg=GREEN, bg=TOTAL_BG, sz=9, bold=True, ha='right')
        c.number_format = CURR

        # I: Running Total
        if i == 0:
            ri_formula = f'=IF(H{row}="","",H{row})'
        else:
            ri_formula = f'=IF(H{row}="","",IF(I{row-1}="",H{row},I{row-1}+H{row}))'
        c = ws.cell(row, 9, ri_formula)
        sty(c, fg=PURPLE, bg=bg, sz=9, bold=True, ha='right')
        c.number_format = CURR

    # Search highlight conditional format
    hl_formula = (f'=AND($C$1<>"",'
                  f'OR(ISNUMBER(SEARCH($C$1,$D{LD_S})),'
                  f'ISNUMBER(SEARCH($C$1,$E{LD_S})),'
                  f'ISNUMBER(SEARCH($C$1,$B{LD_S}))))')
    ds_hl = DifferentialStyle(
        font=Font(name=F, color="1A1A1A", bold=True),
        fill=PatternFill(start_color="FFF176", end_color="FFF176", fill_type='solid')
    )
    ws.conditional_formatting.add(f'A{LD_S}:I{LD_E}',
        Rule(type='expression', formula=[hl_formula], dxf=ds_hl))

    # Print setup
    ws.print_area = f'A1:I{LD_E+5}'
    print_setup(ws, title_rows='1:5')

    print("  ✓ Ledger")


# ══════════════════════════════════════════════════════════════════════════════
# REPORTS SHEET
# ══════════════════════════════════════════════════════════════════════════════
def build_reports(wb):
    ws = wb.create_sheet("Reports")
    ws.sheet_view.showGridLines = False

    col_widths = [16, 14, 18, 8, 12, 8, 16, 14, 10, 18, 14, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Rows 1-2: title
    ws.row_dimensions[1].height = 30
    c = ws.cell(1, 1, "ZAKAT & SADAQAH — PERSON REPORT")
    hdr(c, bg=HEADER_BG, sz=13)
    ws.merge_cells('A1:I1')

    ws.row_dimensions[2].height = 26
    c = ws.cell(2, 1,
        "Select a person from the dropdown in C4. Breakdown cards update automatically. "
        "Use year filter in C5 to narrow by year.")
    sty(c, fg="444444", bg="F0F4F8", sz=9, italic=True, ha='center', wrap=True)
    ws.merge_cells('A2:I2')

    # Row 3: Unique Names helper in col K
    ws.row_dimensions[3].height = 18
    c = ws.cell(3, 11, "Unique Names")
    sty(c, fg="CCCCCC", sz=7, ha='center')
    for i in range(50):
        row = 4 + i
        if i == 0:
            formula = (f'=IFERROR(INDEX(Ledger!$D${LD_S}:$D${LD_E},'
                       f'MATCH(0,COUNTIF($K$3:K3,Ledger!$D${LD_S}:$D${LD_E}),0)),"")')
        else:
            formula = (f'=IFERROR(INDEX(Ledger!$D${LD_S}:$D${LD_E},'
                       f'MATCH(0,COUNTIF($K$3:K{3+i},Ledger!$D${LD_S}:$D${LD_E}),0)),"")')
        ws.cell(row, 11, formula)

    # Row 4: Person selector
    ws.row_dimensions[4].height = 30

    c = ws.cell(4, 1, "Select Person:")
    sty(c, fg=WHITE, bg=SUBHDR_BG, sz=11, bold=True, ha='right')
    ws.merge_cells('A4:B4')

    c = ws.cell(4, 3, "")
    sty(c, fg=BLUE, bg="E8F8F5", sz=12, bold=True, ha='center')
    c.border = th_bdr()

    person_dv = DataValidation(type="list", formula1="$K$4:$K$53",
                                allow_blank=True, showDropDown=False)
    ws.add_data_validation(person_dv)
    person_dv.add(ws['C4'])

    c = ws.cell(4, 4, "Total Given:")
    hdr(c, bg=SUBHDR_BG, sz=10)
    ws.merge_cells('D4:E4')

    c = ws.cell(4, 6,
        f'=IF($C$4="","—",IF($C$5="All Years",'
        f'SUMIF(Ledger!$D${LD_S}:$D${LD_E},$C$4,Ledger!$H${LD_S}:$H${LD_E}),'
        f'SUMPRODUCT((Ledger!$D${LD_S}:$D${LD_E}=$C$4)*'
        f'(YEAR(Ledger!$A${LD_S}:$A${LD_E})=VALUE($C$5))*'
        f'Ledger!$H${LD_S}:$H${LD_E})))')
    sty(c, fg=GREEN, bg=TOTAL_BG, sz=12, bold=True, ha='center')
    c.border = th_bdr(); c.number_format = CURR
    ws.merge_cells('F4:G4')

    c = ws.cell(4, 8, "# Transactions:")
    hdr(c, bg=SUBHDR_BG, sz=10)

    c = ws.cell(4, 9,
        f'=IF($C$4="","—",IF($C$5="All Years",'
        f'COUNTIF(Ledger!$D${LD_S}:$D${LD_E},$C$4),'
        f'SUMPRODUCT((Ledger!$D${LD_S}:$D${LD_E}=$C$4)*'
        f'(YEAR(Ledger!$A${LD_S}:$A${LD_E})=VALUE($C$5)))))')
    sty(c, fg=PURPLE, bg=LED_BG, sz=12, bold=True, ha='center')
    c.border = th_bdr()

    # Row 5: Year filter
    ws.row_dimensions[5].height = 28

    c = ws.cell(5, 1, "Filter by Year:")
    sty(c, fg=WHITE, bg=SUBHDR_BG, sz=11, bold=True, ha='right')
    ws.merge_cells('A5:B5')

    c = ws.cell(5, 3, "All Years")
    sty(c, fg=BLUE, bg="E8F8F5", sz=11, bold=True, ha='center')
    c.border = th_bdr()

    # Year list in col M
    ws.column_dimensions['M'].width = 10
    ws.cell(5, 13, "All Years")
    for i in range(1, 20):
        ws.cell(5+i, 13, "")

    yr_dv = DataValidation(type="list", formula1="$M$5:$M$24",
                            allow_blank=True, showDropDown=False)
    ws.add_data_validation(yr_dv)
    yr_dv.add(ws['C5'])

    c = ws.cell(5, 4, "Select a year to filter transactions, or 'All Years' to see everything")
    sty(c, fg="555555", bg="F8F9FA", sz=9, italic=True)
    ws.merge_cells('D5:I5')

    # Row 6: Breakdown header
    ws.row_dimensions[6].height = 24
    c = ws.cell(6, 1, "BREAKDOWN BY TYPE  (auto-updates when you add types in Settings)")
    hdr(c, bg=HEADER_BG, sz=10)
    ws.merge_cells('A6:I6')

    # Row 7-8: Sub-headers
    ws.row_dimensions[7].height = 22
    c = ws.cell(7, 1, "Type")
    hdr(c, bg=SUBHDR_BG, sz=9)
    ws.merge_cells('A7:E7')
    c = ws.cell(7, 6, "Total Paid to Selected Person (Amount + Fees)")
    hdr(c, bg=SUBHDR_BG, sz=9)
    ws.merge_cells('F7:I7')

    # Type breakdown rows 8-37 (30 types)
    for i in range(30):
        row = 8 + i
        ws.row_dimensions[row].height = 20
        bg = "F0F4F8" if i % 2 == 0 else WHITE

        c = ws.cell(row, 1, f'=IF(Settings!$B${5+i}="","",Settings!$B${5+i})')
        sty(c, fg=BLACK, bg=bg, sz=9, bold=True)
        ws.merge_cells(f'A{row}:E{row}')

        c = ws.cell(row, 6,
            f'=IF(Settings!$B${5+i}="","",IF($C$4="","—",'
            f'SUMPRODUCT((Ledger!$D${LD_S}:$D${LD_E}=$C$4)*'
            f'(Ledger!$B${LD_S}:$B${LD_E}=Settings!$B${5+i})*'
            f'IF($C$5="All Years",1,YEAR(Ledger!$A${LD_S}:$A${LD_E})=IF($C$5="All Years",0,VALUE($C$5)))*'
            f'Ledger!$H${LD_S}:$H${LD_E})))')
        sty(c, fg=GREEN, bg=bg, sz=9, ha='right')
        c.number_format = CURR
        ws.merge_cells(f'F{row}:I{row}')

    # Rows 38-40: spacer + total
    TOTAL_ROW = 39
    ws.row_dimensions[TOTAL_ROW].height = 24
    c = ws.cell(TOTAL_ROW, 1, "TOTAL (all types)")
    hdr(c, bg=SUBHDR_BG, sz=9)
    ws.merge_cells(f'A{TOTAL_ROW}:E{TOTAL_ROW}')
    c = ws.cell(TOTAL_ROW, 6,
        f'=IF($C$4="","—",SUMPRODUCT((Ledger!$D${LD_S}:$D${LD_E}=$C$4)*'
        f'IF($C$5="All Years",1,YEAR(Ledger!$A${LD_S}:$A${LD_E})=IF($C$5="All Years",0,VALUE($C$5)))*'
        f'Ledger!$H${LD_S}:$H${LD_E}))')
    sty(c, fg=GREEN, bg=TOTAL_BG, sz=10, bold=True, ha='right')
    c.number_format = CURR; c.border = th_bdr()
    ws.merge_cells(f'F{TOTAL_ROW}:I{TOTAL_ROW}')

    # Row 41: Detail table header
    TBL_HDR = 41
    ws.row_dimensions[TBL_HDR].height = 22
    for col, lbl in enumerate(["#","Date","Type","Service Used","Given To",
                                "Details / Notes","Amount ($)","Fees ($)","Total Paid ($)"], 1):
        c = ws.cell(TBL_HDR, col, lbl)
        hdr(c, bg=SUBHDR_BG, sz=9)

    ws.freeze_panes = 'A11'

    # L helpers rows 41-140 (n=1 at L41), detail data rows 42-141
    TBL_DATA_S = 42;  TBL_DATA_E = 141
    for i in range(100):
        n   = i + 1
        l_r = 41 + i   # L41=n1, L42=n2 ...
        formula = (
            f'=IFERROR(SMALL('
            f'IF((Ledger!$D${LD_S}:$D${LD_E}=$C$4)*'
            f'IF($C$5="All Years",1,YEAR(Ledger!$A${LD_S}:$A${LD_E})='
            f'IF($C$5="All Years",0,VALUE($C$5))),'
            f'ROW(Ledger!$D${LD_S}:$D${LD_E})-ROW(Ledger!$D${LD_S})+1),{n}),"")'
        )
        ws.cell(l_r, 12, formula)

    for i in range(100):
        data_row = TBL_DATA_S + i
        l_ref    = 41 + i
        n        = i + 1
        ws.row_dimensions[data_row].height = 20
        bg = "F8F9FA" if i % 2 == 0 else WHITE

        force(ws, data_row, 1, f'=IF($L{l_ref}="","",{n})')
        sty(ws.cell(data_row, 1), fg="888888", bg=bg, sz=8, ha='center')

        for col, led_col in [(2,'A'),(3,'B'),(4,'C'),(5,'D'),(6,'E'),(7,'F'),(8,'G'),(9,'H')]:
            formula = (
                f'=IFERROR(IF($L{l_ref}="","",INDEX('
                f'Ledger!${led_col}${LD_S}:${led_col}${LD_E},$L{l_ref})),"—")'
            )
            cell = force(ws, data_row, col, formula)
            cell_sty_fg = BLACK
            if col == 2: cell.number_format = DATEFMT
            if col in [7, 8, 9]: cell.number_format = CURR; cell_sty_fg = GREEN
            sty(cell, fg=cell_sty_fg, bg=bg, sz=9)

    # Service fee summary
    SVC_TITLE = TBL_DATA_E + 3
    ws.row_dimensions[SVC_TITLE].height = 26
    c = ws.cell(SVC_TITLE, 1,
        "FEES PAID BY SERVICE — All Time (independent of person / year filter)")
    hdr(c, bg=HEADER_BG, sz=10)
    ws.merge_cells(f'A{SVC_TITLE}:I{SVC_TITLE}')

    SVC_HDR = SVC_TITLE + 1
    ws.row_dimensions[SVC_HDR].height = 22
    for col, end, txt in [(1,3,"Service"),(4,5,"Total Amount ($)"),(6,7,"Total Fees ($)"),(8,9,"# Payments")]:
        c = ws.cell(SVC_HDR, col, txt)
        hdr(c, bg=SUBHDR_BG, sz=9)
        ws.merge_cells(f'{get_column_letter(col)}{SVC_HDR}:{get_column_letter(end)}{SVC_HDR}')

    for i in range(30):
        row = SVC_HDR + 1 + i
        ws.row_dimensions[row].height = 20
        bg = "F8F9FA" if i % 2 == 0 else WHITE
        sc = f"Settings!$D${5+i}"

        c = ws.cell(row, 1, f'=IF({sc}="","",{sc})')
        sty(c, fg=BLACK, bg=bg, sz=9, bold=True)
        ws.merge_cells(f'A{row}:C{row}')

        c = ws.cell(row, 4,
            f'=IF({sc}="","",SUMIF(Ledger!$C${LD_S}:$C${LD_E},{sc},Ledger!$F${LD_S}:$F${LD_E}))')
        sty(c, fg=GREEN, bg=bg, sz=9, ha='right'); c.number_format = CURR
        ws.merge_cells(f'D{row}:E{row}')

        c = ws.cell(row, 6,
            f'=IF({sc}="","",SUMIF(Ledger!$C${LD_S}:$C${LD_E},{sc},Ledger!$G${LD_S}:$G${LD_E}))')
        sty(c, fg=RED, bg=bg, sz=9, ha='right'); c.number_format = CURR
        ws.merge_cells(f'F{row}:G{row}')

        c = ws.cell(row, 8,
            f'=IF({sc}="","",COUNTIF(Ledger!$C${LD_S}:$C${LD_E},{sc}))')
        sty(c, fg=PURPLE, bg=bg, sz=9, ha='center')
        ws.merge_cells(f'H{row}:I{row}')

    # Print setup
    ws.print_area = f'A1:I{SVC_HDR+31}'
    print_setup(ws, orient='portrait', title_rows='1:5')

    print("  ✓ Reports")


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════
def main():
    print("Building Zakat Tracker...")

    wb = Workbook()
    wb.remove(wb.active)   # remove default blank sheet

    build_guide(wb)
    build_settings(wb)
    build_summary(wb)

    build_asset_sheet(wb, 'Stocks',
        ["TD Ameritrade","Charles Schwab","Fidelity","Vanguard","Robinhood","Other Account"],
        "Total Portfolio")
    build_asset_sheet(wb, 'Cash',
        ["Chase Checking","Chase Savings","Bank of America","Money Market","Cash on Hand","Other Liquid"],
        "Total Cash")
    build_asset_sheet(wb, 'Debts',
        ["Chase Credit Card","Citi Credit Card","Amex Credit Card","Car Loan","Personal Loan","Other Debt"],
        "Total Debts")

    build_ledger(wb)
    build_reports(wb)

    # Font sweep — ensure all cells use Aptos
    for sname in wb.sheetnames:
        for row in wb[sname].iter_rows():
            for cell in row:
                if cell.font and cell.font.name != F:
                    old = cell.font
                    cell.font = Font(name=F, bold=old.bold, italic=old.italic,
                                     color=old.color, size=old.size,
                                     underline=old.underline, strike=old.strike)

    out = 'Zakat_Tracker.xlsx'
    wb.save(out)
    print(f"\n✅  Saved: {out}")
    print(f"   Sheets: {', '.join(wb.sheetnames)}")


if __name__ == '__main__':
    main()
